import time
import shutil
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook

# === File Paths ===
SOURCE_FILE = "calculator_test_cases.xlsx"
RESULT_FILE = "calculator_test_cases_result.xlsx"

# === Step 1: Copy source file to result file ===
if os.path.exists(RESULT_FILE):
    os.remove(RESULT_FILE)
shutil.copy(SOURCE_FILE, RESULT_FILE)

# === Step 2: Load workbook and select the correct sheet ===
wb = load_workbook(RESULT_FILE)
sheet = wb.active  # If your sheet has a name, use wb['SheetName']

# === Step 3: Launch the browser ===
driver = webdriver.Chrome(service=Service())
driver.get("http://localhost:5000")
time.sleep(1)

# === Step 4: Iterate through test cases ===
for row in sheet.iter_rows(min_row=2):  # Include cell objects
    input_a = row[2].value  # Column C
    operation = row[3].value  # Column D
    input_b = row[4].value  # Column E
    expected = row[5].value  # Column F

    # Input values into the web form
    driver.find_element(By.NAME, "num1").clear()
    driver.find_element(By.NAME, "num1").send_keys(str(input_a))
    driver.find_element(By.NAME, "num2").clear()
    driver.find_element(By.NAME, "num2").send_keys(str(input_b))
    Select(driver.find_element(By.NAME, "operation")).select_by_value(operation)
    driver.find_element(By.XPATH, "//button[@type='submit']").click()
    time.sleep(1)

    # Capture result or error
    try:
        result_elem = driver.find_element(By.XPATH, "//p[strong[text()='Result:']]")
        actual_result = result_elem.text.split(":")[1].strip()
    except:
        try:
            error_elem = driver.find_element(By.XPATH, "//p[strong[text()='Error:']]")
            actual_result = error_elem.text.split(":")[1].strip()
        except:
            actual_result = "Unknown"

    # === Write back to Excel ===
    excel_row = row[0].row  # Get the actual Excel row number

    # Write to Column G (Actual) and H (Status)
    sheet.cell(row=excel_row, column=7, value=actual_result)

    try:
        # Numerical comparison
        if float(expected) == float(actual_result):
            status = "Pass"
        else:
            status = "Fail"
    except:
        # Text comparison
        if str(expected).strip().lower() == str(actual_result).strip().lower():
            status = "Pass"
        else:
            status = "Fail"

    sheet.cell(row=excel_row, column=8, value=status)

    # Optional debug output
    print(f"[{excel_row}] {input_a} {operation} {input_b} → {actual_result} (Expected: {expected}) = {status}")

# === Step 5: Save and close ===
wb.save(RESULT_FILE)
driver.quit()

print(f"✅ Results written to: {os.path.abspath(RESULT_FILE)}")
